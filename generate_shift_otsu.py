"""大津京こと薬局 シフト表生成スクリプト
起動時に翌月を自動設定。

【固定パターン（コード内・毎月変わらない）】
  B: 月=遅番(10-19), 火=遅番(10-19), 木=早番(9-18)  水固定休
  C: 月=遅番(10-19), 水=9-17,         木=遅番(10-19)  火固定休
  D: 火=遅番(10-19), 水=9-17,         木=遅番(10-19)  月固定休
  金: 出勤者は遅番(10-19)。b_fri/c_fri/d_fri に出勤日を指定（2人出勤）
  土: 出勤者は9-17。b_sat/c_sat/d_sat に出勤日を指定（2人出勤）
  日祝: 1人ローテーション
  H: 月=遅番, 火=早番, 第1,2水=8:45-17, 木=遅番, 金=早番, 土=条件付き
  I: 月=早番, 火=遅番, 第3,4水=8:45-17, 木=早番, 金=遅番, 土=条件付き
  J: 月=早番, 火=遅番, 木=遅番, 金=早番, 土=9-17  (水固定休)

【日祝自動ローテーション】
  薬剤師: B→C→D→B→... の順で自動割り当て
  事務  : H→I→J→H→... の順で自動割り当て
  日曜担当者は同週の土曜をオフにして週休2日を維持
  上書きしたい場合: sun_override_ph / sun_override_jm に {日付:"X"} を記入

【毎月更新が必要な変数】
  HOLIDAYS       : 祝日set（日付番号）
  sun_override_ph: 日祝薬剤師担当の上書き（空欄=自動）
  sun_override_jm: 日祝事務担当の上書き（空欄=自動）
  *_extra_off    : 追加休日（自動パターンから除外したい日）
  requested_off  : 希望休（備考欄に「×」）
  yukyu_per_person: 有休（備考欄に「有休」）
  TARGETS        : 目標時間（B,C,D,H,I,J → 毎月設定）

【シフト時間】
  早番(B/C/D) : 9:00-18:00 (8h net)    遅番(B/C/D) : 10:00-19:00 (8h net)
  早番(H/I/J) : 8:45-18:00 (8.25h net) 遅番(H/I/J) : 10:00-19:00 (8h net)
  水/土/日祝  : 9:00-17:00 (7h net)
"""

import os
import datetime
import calendar
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins

# ================================================================
# 翌月自動設定
# ================================================================
_today = datetime.date.today()
YEAR  = _today.year + 1 if _today.month == 12 else _today.year
MONTH = 1 if _today.month == 12 else _today.month + 1

DAYS_IN_MONTH = calendar.monthrange(YEAR, MONTH)[1]
WEEKDAY_NAMES = {0:"月", 1:"火", 2:"水", 3:"木", 4:"金", 5:"土", 6:"日"}

# ================================================================
# 毎月更新
# ================================================================
HOLIDAYS = set()   # 臨時休業日（祝日は通常営業）

# 日祝担当の上書き（空欄=自動ローテーション）
# 例: sun_override_ph = {5: "C"}  → 5日だけCに変更
sun_override_ph = {}
sun_override_jm = {}

# 金曜出勤日（遅番 10:00-19:00）2人出勤になるよう指定
b_fri = set()   # 例: {3, 10, 17, 24, 31}
c_fri = set()
d_fri = set()

# 土曜出勤日（9:00-17:00）2人出勤になるよう指定
b_sat = set()   # 例: {4, 11, 18, 25}
c_sat = set()
d_sat = set()

# 追加休日（自動パターンから除外したい日）
b_extra_off = set()
c_extra_off = set()
d_extra_off = set()
h_extra_off = set()
i_extra_off = set()
j_extra_off = set()

# 希望休（備考欄に「×」）
requested_off = {k: set() for k in "ABCDEFGHIJK"}

# 有休（備考欄に「有休」）
yukyu_per_person = {k: set() for k in "ABCDEFGHIJK"}

# 目標時間（A,E,F,G は自動計算。B,C,D,H,I,J は毎月設定）
TARGETS = {
    "A": None,  # 自動計算
    "B": None,  # 正社員: 要設定 例 173.0
    "C": None,
    "D": None,
    "E": None,  # 自動計算
    "F": None,  # 自動計算
    "G": None,  # 自動計算
    "H": None,  # 正社員: 要設定
    "I": None,
    "J": None,
    "K": None,
}

# ================================================================
# スタッフ名
# ================================================================
NAMES = {
    "A": "宮﨑 恵未",  "B": "中本 貴士",   "C": "今堀 翔太",
    "D": "安井 桜",    "E": "石田 真澄",   "F": "筒井 麻耶",
    "G": "黒田 梨恵",  "H": "西山 美紗子", "I": "原田 真衣",
    "J": "和田 美聖",  "K": "小寺 和彦",
}
PERSONS = list("ABCDEFGHIJK")

# 会社指定 正社員 所定労働時間（年間固定）
SHOTEIKOJI = {
    2026: {1:151, 2:143, 3:169, 4:163, 5:153,
           6:171, 7:169, 8:169, 9:147, 10:169, 11:155, 12:165},
}

# ================================================================
# ヘルパー
# ================================================================
def is_sun_hol(d):
    return datetime.date(YEAR, MONTH, d).weekday() == 6 or d in HOLIDAYS

def net_hours(s, e):
    h1,m1 = map(int,s.split(":")); h2,m2 = map(int,e.split(":"))
    g = (h2*60+m2 - h1*60-m1)/60
    return g-1.0 if g>=7 else (g-0.5 if g>6 else g)

def skip(name, d):
    return d in requested_off[name] or d in yukyu_per_person[name]

# 第1,2水曜 → H担当 / 第3水曜以降 → I担当
_weds = [d for d in range(1, DAYS_IN_MONTH+1)
         if datetime.date(YEAR,MONTH,d).weekday() == 2]
h_wed = set(_weds[:2])
i_wed = set(_weds[2:])

def sat_has_wed(sat_d, wed_set):
    """土曜と同じ週の水曜が wed_set に含まれるか（Sat-3 = 同週Wed）"""
    wed_day = sat_d - 3
    return 1 <= wed_day <= DAYS_IN_MONTH and wed_day in wed_set

# ================================================================
# 日祝ローテーション自動計算
# ================================================================
# 固定休曜日（その曜日の祝日はその人を担当から外す）
PH_FIXED_OFF_WD = {"B": 2, "C": 1, "D": 0}   # B=水, C=火, D=月

sun_hol_days = sorted(d for d in range(1, DAYS_IN_MONTH+1) if is_sun_hol(d))

ph_rot = ["B", "C", "D"]
jm_rot = ["H", "I", "J"]

# 自動ローテーション（固定休曜日の祝日はスキップ）
ph_assign = {}   # {日付: 担当薬剤師キー}
jm_assign = {}   # {日付: 担当事務キー}

ph_idx = 0
jm_idx = 0
for d in sun_hol_days:
    wd = datetime.date(YEAR, MONTH, d).weekday()
    # 薬剤師: 固定休が一致する人はスキップ
    if d in sun_override_ph:
        ph_assign[d] = sun_override_ph[d]
    else:
        for offset in range(3):
            candidate = ph_rot[(ph_idx + offset) % 3]
            if PH_FIXED_OFF_WD.get(candidate) != wd:
                ph_assign[d] = candidate
                ph_idx = (ph_idx + offset + 1) % 3
                break
        else:
            ph_assign[d] = ph_rot[ph_idx % 3]
            ph_idx = (ph_idx + 1) % 3
    # 事務: 固定休なし、単純ローテーション
    if d in sun_override_jm:
        jm_assign[d] = sun_override_jm[d]
    else:
        jm_assign[d] = jm_rot[jm_idx % 3]
        jm_idx = (jm_idx + 1) % 3

# 日曜出勤の補償日（週休2日維持・6連勤防止）
#
# B/C/D/J: 前日土曜オフ（日曜の前日）
#   → B は水曜固定休のため「木金[土off]日月火[水off]」= 最大3連勤
#   → B/C は月曜が出勤のまま維持 → 月曜は B+C=2人確保
#
# H/I: 翌月曜オフ
#   → 水曜シフトが続くため土曜補償だと日〜金で6連勤になるため
#
# ※祝日(平日)は補償なし（別途 *_extra_off で調整）
extra_off = {"B": b_extra_off, "C": c_extra_off, "D": d_extra_off,
             "H": h_extra_off, "I": i_extra_off, "J": j_extra_off}

for d, person in list(ph_assign.items()) + list(jm_assign.items()):
    if datetime.date(YEAR, MONTH, d).weekday() == 6:   # 日曜のみ
        # H, I : 翌月曜補償（水曜シフト連続で6連勤防止）
        # B,C,D,J: 前日土曜補償（週休2日維持・6連勤防止）
        comp_d = (d + 1) if person in ("H", "I") else (d - 1)
        if 1 <= comp_d <= DAYS_IN_MONTH and person in extra_off:
            extra_off[person].add(comp_d)

# H：第1・2水曜の週の月曜を自動休みに
for d in range(1, DAYS_IN_MONTH + 1):
    if datetime.date(YEAR, MONTH, d).weekday() == 0:
        wed_of_week = d + 2
        if wed_of_week <= DAYS_IN_MONTH and wed_of_week in h_wed:
            extra_off["H"].add(d)

# ================================================================
# 金・土 自動配置（b_fri/c_fri/d_fri/b_sat/c_sat/d_sat が未設定の場合のみ）
# 「所定に最も近くなるよう」残り必要時間が多い2人を各日に割り当てる
# 土曜: 日曜duty の補償日はスキップ（既に extra_off 済み）
# 金曜: B/C/D 全員対象（6連勤にならないことを確認済み）
# ================================================================
if not any([b_fri, c_fri, d_fri, b_sat, c_sat, d_sat]):
    _shoteikoji = SHOTEIKOJI.get(YEAR, {}).get(MONTH, 0)
    if _shoteikoji:
        # ベース時間（月火木・水 + 日祝duty）を推計
        _base = {"B": 0.0, "C": 0.0, "D": 0.0}
        for _d in range(1, DAYS_IN_MONTH + 1):
            _wd = datetime.date(YEAR, MONTH, _d).weekday()
            _sh = is_sun_hol(_d)
            for _p in ("B", "C", "D"):
                if _sh:
                    if ph_assign.get(_d) == _p:
                        _base[_p] += 7.0
                else:
                    if _p == "B" and _wd not in (2, 5, 4):  # 月火木のみ
                        if _wd in (0, 1, 3): _base["B"] += 8.0
                    elif _p == "C" and _wd not in (1, 5, 4):  # 月水木のみ
                        if _wd in (0, 3):    _base["C"] += 8.0
                        elif _wd == 2:       _base["C"] += 7.0
                    elif _p == "D" and _wd not in (0, 5, 4):  # 火水木のみ
                        if _wd in (1, 3):    _base["D"] += 8.0
                        elif _wd == 2:       _base["D"] += 7.0

        _rem  = {_p: max(0.0, _shoteikoji - _base[_p]) for _p in ("B","C","D")}
        _comp = {"B": b_extra_off, "C": c_extra_off, "D": d_extra_off}
        _fset = {"B": b_fri, "C": c_fri, "D": d_fri}
        _sset = {"B": b_sat, "C": c_sat, "D": d_sat}

        # 土曜: 補償日を除いた中から残り必要時間が多い2人
        for _d in sorted(d for d in range(1, DAYS_IN_MONTH+1)
                         if datetime.date(YEAR,MONTH,d).weekday()==5 and not is_sun_hol(d)):
            _avail = [_p for _p in ("B","C","D") if _d not in _comp[_p]]
            for _p in sorted(_avail, key=lambda _p: -_rem[_p])[:2]:
                _sset[_p].add(_d); _rem[_p] -= 7.0

        # 金曜: 残り必要時間が多い2人（連勤は最大5日で問題なし）
        for _d in sorted(d for d in range(1, DAYS_IN_MONTH+1)
                         if datetime.date(YEAR,MONTH,d).weekday()==4 and not is_sun_hol(d)):
            for _p in sorted(("B","C","D"), key=lambda _p: -_rem[_p])[:2]:
                _fset[_p].add(_d); _rem[_p] -= 8.0

        print(f"[自動配置] 金: B={sorted(b_fri)} C={sorted(c_fri)} D={sorted(d_fri)}")
        print(f"[自動配置] 土: B={sorted(b_sat)} C={sorted(c_sat)} D={sorted(d_sat)}")

# ================================================================
# シフトデータ生成
# ================================================================
shift_data = {n:{} for n in PERSONS}

for d in range(1, DAYS_IN_MONTH+1):
    dt = datetime.date(YEAR, MONTH, d)
    wd = dt.weekday()
    sh = is_sun_hol(d)

    # --- A: 月火水木金(祝日除く) ---
    if not skip("A",d) and not sh:
        if   wd == 0: shift_data["A"][d] = ("8:45","17:30", net_hours("8:45","17:30"), False)
        elif wd == 1: shift_data["A"][d] = ("8:45","17:30", net_hours("8:45","17:30"), False)
        elif wd == 2: shift_data["A"][d] = ("9:00","17:00", net_hours("9:00","17:00"),  False)
        elif wd == 3: shift_data["A"][d] = ("8:45","17:30", net_hours("8:45","17:30"), False)
        elif wd == 4: shift_data["A"][d] = ("8:45","17:30", net_hours("8:45","17:30"), False)

    # --- B: 水固定休 / 月=遅, 火=遅, 木=早, 金=遅(b_fri指定日), 土=9-17(b_sat指定日) ---
    if not skip("B",d) and wd != 2 and d not in b_extra_off:
        if sh:
            if ph_assign.get(d) == "B":
                shift_data["B"][d] = ("9:00","17:00", net_hours("9:00","17:00"), False)
        elif wd == 0: shift_data["B"][d] = ("10:00","19:00", net_hours("10:00","19:00"), False)
        elif wd == 1: shift_data["B"][d] = ("10:00","19:00", net_hours("10:00","19:00"), False)
        elif wd == 3: shift_data["B"][d] = ("9:00", "18:00", net_hours("9:00","18:00"),  False)
        elif wd == 4 and d in b_fri: shift_data["B"][d] = ("10:00","19:00", net_hours("10:00","19:00"), False)
        elif wd == 5 and d in b_sat: shift_data["B"][d] = ("9:00", "17:00", net_hours("9:00","17:00"),  False)

    # --- C: 火固定休 / 月=遅, 水=9-17, 木=遅, 金=遅(c_fri指定日), 土=9-17(c_sat指定日) ---
    if not skip("C",d) and wd != 1 and d not in c_extra_off:
        if sh:
            if ph_assign.get(d) == "C":
                shift_data["C"][d] = ("9:00","17:00", net_hours("9:00","17:00"), False)
        elif wd == 0: shift_data["C"][d] = ("10:00","19:00", net_hours("10:00","19:00"), False)
        elif wd == 2: shift_data["C"][d] = ("9:00", "17:00", net_hours("9:00","17:00"),  False)
        elif wd == 3: shift_data["C"][d] = ("10:00","19:00", net_hours("10:00","19:00"), False)
        elif wd == 4 and d in c_fri: shift_data["C"][d] = ("10:00","19:00", net_hours("10:00","19:00"), False)
        elif wd == 5 and d in c_sat: shift_data["C"][d] = ("9:00", "17:00", net_hours("9:00","17:00"),  False)

    # --- D: 月固定休 / 火=遅, 水=9-17, 木=遅, 金=遅(d_fri指定日), 土=9-17(d_sat指定日) ---
    if not skip("D",d) and wd != 0 and d not in d_extra_off:
        if sh:
            if ph_assign.get(d) == "D":
                shift_data["D"][d] = ("9:00","17:00", net_hours("9:00","17:00"), False)
        elif wd == 1: shift_data["D"][d] = ("10:00","19:00", net_hours("10:00","19:00"), False)
        elif wd == 2: shift_data["D"][d] = ("9:00", "17:00", net_hours("9:00","17:00"),  False)
        elif wd == 3: shift_data["D"][d] = ("10:00","19:00", net_hours("10:00","19:00"), False)
        elif wd == 4 and d in d_fri: shift_data["D"][d] = ("10:00","19:00", net_hours("10:00","19:00"), False)
        elif wd == 5 and d in d_sat: shift_data["D"][d] = ("9:00", "17:00", net_hours("9:00","17:00"),  False)

    # --- E: 月火木金 9:00-13:00 (祝日除く) ---
    if not skip("E",d) and not sh:
        if wd in (0,1,3,4):
            shift_data["E"][d] = ("9:00","13:00", net_hours("9:00","13:00"), False)

    # --- F: 月火金 9:00-14:00 (祝日除く) ---
    if not skip("F",d) and not sh:
        if wd in (0,1,4):
            shift_data["F"][d] = ("9:00","14:00", net_hours("9:00","14:00"), False)

    # --- G: 木 9:00-13:00 / 土 9:00-17:00 ---
    if not skip("G",d):
        if wd == 3 and not sh:
            shift_data["G"][d] = ("9:00","13:00", net_hours("9:00","13:00"), False)
        elif wd == 5:
            shift_data["G"][d] = ("9:00","17:00", net_hours("9:00","17:00"), False)

    # --- H: 月=早, 火=遅, 第1,2水=9-17, 木=遅, 金=早, 土=条件付き ---
    if not skip("H",d) and d not in h_extra_off:
        if sh:
            if jm_assign.get(d) == "H":
                shift_data["H"][d] = ("9:00","17:00", net_hours("9:00","17:00"), False)
        elif d in h_wed:
            shift_data["H"][d] = ("9:00","17:00", net_hours("9:00","17:00"), False)
        elif wd == 0: shift_data["H"][d] = ("10:00","19:00", net_hours("10:00","19:00"), False)
        elif wd == 1: shift_data["H"][d] = ("8:45","18:00",  net_hours("8:45","18:00"),  False)
        elif wd == 3: shift_data["H"][d] = ("10:00","19:00", net_hours("10:00","19:00"), False)
        elif wd == 4: shift_data["H"][d] = ("8:45","18:00",  net_hours("8:45","18:00"),  False)
        elif wd == 5 and not sat_has_wed(d, h_wed):
            shift_data["H"][d] = ("8:45","17:00", net_hours("8:45","17:00"), False)

    # --- I: 月=早, 火=遅, 第3,4水=9-17, 木=遅, 金=早, 土=条件付き ---
    if not skip("I",d) and d not in i_extra_off:
        if sh:
            if jm_assign.get(d) == "I":
                shift_data["I"][d] = ("9:00","17:00", net_hours("9:00","17:00"), False)
        elif d in i_wed:
            shift_data["I"][d] = ("9:00","17:00", net_hours("9:00","17:00"), False)
        elif wd == 0: shift_data["I"][d] = ("8:45","18:00",  net_hours("8:45","18:00"),  False)
        elif wd == 1: shift_data["I"][d] = ("10:00","19:00", net_hours("10:00","19:00"), False)
        elif wd == 3: shift_data["I"][d] = ("8:45","18:00",  net_hours("8:45","18:00"),  False)
        elif wd == 4: shift_data["I"][d] = ("10:00","19:00", net_hours("10:00","19:00"), False)
        elif wd == 5 and not sat_has_wed(d, i_wed):
            shift_data["I"][d] = ("8:45","17:00", net_hours("8:45","17:00"), False)

    # --- J: 月=早, 火=遅, 木=早, 金=遅, 土=9-17, 水・日=休 ---
    if not skip("J",d) and wd != 2 and d not in j_extra_off:
        if sh:
            if jm_assign.get(d) == "J":
                shift_data["J"][d] = ("9:00","17:00", net_hours("9:00","17:00"), False)
        elif wd == 0: shift_data["J"][d] = ("8:45","18:00",  net_hours("8:45","18:00"),  False)
        elif wd == 1: shift_data["J"][d] = ("10:00","19:00", net_hours("10:00","19:00"), False)
        elif wd == 3: shift_data["J"][d] = ("10:00","19:00", net_hours("10:00","19:00"), False)
        elif wd == 4: shift_data["J"][d] = ("8:45","18:00",  net_hours("8:45","18:00"),  False)
        elif wd == 5: shift_data["J"][d] = ("8:45","17:00",  net_hours("8:45","17:00"),  False)

    # J カバー調整: H/I どちらかが不在で早晩どちらかが欠ける場合に J が補完
    if not sh and wd in (0, 1, 3, 4) and d in shift_data["J"]:
        h_st = shift_data["H"].get(d, (None,))[0]
        i_st = shift_data["I"].get(d, (None,))[0]
        has_early = h_st == "8:45" or i_st == "8:45"
        has_late  = h_st == "10:00" or i_st == "10:00"
        if has_early and not has_late:
            shift_data["J"][d] = ("10:00","19:00", net_hours("10:00","19:00"), False)
        elif has_late and not has_early:
            shift_data["J"][d] = ("8:45","18:00",  net_hours("8:45","18:00"),  False)

    # K: 空欄（手動記入）

# E/F/G/K: 実シフト時間から自動計算
for n in ("E","F","G"):
    TARGETS[n] = round(sum(v[2] for v in shift_data[n].values()), 2)

_shoteikoji = SHOTEIKOJI.get(YEAR, {}).get(MONTH)

# A: 所定労働時間 × 0.8
TARGETS["A"] = round(_shoteikoji * 0.8, 2) if _shoteikoji else round(sum(v[2] for v in shift_data["A"].values()), 2)

for n in ("B","C","D","H","I","J"):
    if TARGETS[n] is None:
        TARGETS[n] = _shoteikoji  # None の場合は年間テーブル未登録

# ================================================================
# 時間調整（所定労働時間に近づける）
# ================================================================
def _adjust_hours(shift_data, shoteikoji_val):
    """過剰 → 金土シフトを削除（薬剤師カバレッジ ≥2 を維持）
    不足 → 未出勤の金土にシフトを追加"""

    _extra = {"B": b_extra_off, "C": c_extra_off, "D": d_extra_off,
              "H": h_extra_off, "I": i_extra_off, "J": j_extra_off}

    def _is_sh(d):
        return datetime.date(YEAR, MONTH, d).weekday() == 6 or d in HOLIDAYS

    def _sat_hw(sat_d, wed_set):
        wd = sat_d - 3
        return 1 <= wd <= DAYS_IN_MONTH and wd in wed_set

    for person in ("B", "C", "D", "H", "I", "J"):
        data = shift_data[person]
        if not data:
            continue
        total = sum(v[2] for v in data.values())
        diff = total - shoteikoji_val
        if abs(diff) < 0.5:
            continue
        over = diff > 0

        if person in ("B", "C", "D"):
            if over:
                for wd_t in (5, 4):
                    removable = sorted(
                        [d for d in list(data)
                         if datetime.date(YEAR, MONTH, d).weekday() == wd_t and not _is_sh(d)],
                        reverse=True,
                    )
                    for d in removable:
                        others = [p for p in ("B", "C", "D") if p != person and d in shift_data[p]]
                        if len(others) < 2:
                            continue
                        saved = data[d][2]
                        del shift_data[person][d]
                        diff -= saved
                        if diff <= 0.5:
                            break
                    if diff <= 0.5:
                        break
            else:
                for wd_t, s, e in ((5, "9:00", "17:00"), (4, "10:00", "19:00")):
                    for d in range(1, DAYS_IN_MONTH + 1):
                        if datetime.date(YEAR, MONTH, d).weekday() != wd_t:
                            continue
                        if _is_sh(d):
                            continue
                        if d in shift_data[person]:
                            continue
                        if d in _extra.get(person, set()):
                            continue
                        nh = net_hours(s, e)
                        shift_data[person][d] = (s, e, nh, False)
                        diff += nh
                        if diff >= -0.5:
                            break
                    if diff >= -0.5:
                        break

        elif person in ("H", "I"):
            wed_set = h_wed if person == "H" else i_wed
            if over:
                sat_days = sorted(
                    [d for d in list(data)
                     if datetime.date(YEAR, MONTH, d).weekday() == 5 and not _is_sh(d)],
                    reverse=True,
                )
                for d in sat_days:
                    saved = data[d][2]
                    del shift_data[person][d]
                    diff -= saved
                    if diff <= 0.5:
                        break
            else:
                for d in range(1, DAYS_IN_MONTH + 1):
                    if datetime.date(YEAR, MONTH, d).weekday() != 5:
                        continue
                    if _is_sh(d):
                        continue
                    if d in shift_data[person]:
                        continue
                    if d in _extra.get(person, set()):
                        continue
                    if _sat_hw(d, wed_set):
                        nh = net_hours("8:45", "17:00")
                        shift_data[person][d] = ("8:45", "17:00", nh, False)
                        diff += nh
                        if diff >= -0.5:
                            break

        else:  # J
            if over:
                sat_days = sorted(
                    [d for d in list(data)
                     if datetime.date(YEAR, MONTH, d).weekday() == 5 and not _is_sh(d)],
                    reverse=True,
                )
                for d in sat_days:
                    saved = data[d][2]
                    del shift_data[person][d]
                    diff -= saved
                    if diff <= 0.5:
                        break

if _shoteikoji:
    _adjust_hours(shift_data, _shoteikoji)

# ================================================================
# 色・スタイル
# ================================================================
COLOR_PH    = "4BACC6"
COLOR_JM    = "F79646"
COLOR_OT    = "70AD47"
COLOR_SAT   = "DDEEFF"
COLOR_HOL   = "F2DCDB"
COLOR_LABEL = "F2F2F2"
COLOR_YEL   = "FFFF00"

wb = Workbook()
ws = wb.active
ws.title = "薬局シフト表"
ws.page_setup.paperSize   = ws.PAPERSIZE_A4
ws.page_setup.fitToPage   = True
ws.page_setup.fitToWidth  = 1
ws.page_setup.fitToHeight = 0
ws.page_margins = PageMargins(left=0.7,right=0.7,top=0.75,bottom=0.75,header=0.3,footer=0.3)

COL_STRIPE=1; COL_NAME=2; COL_TYPE=3
DAY_COL_START=4
TOTAL_COL=DAY_COL_START+DAYS_IN_MONTH
TARGET_COL=TOTAL_COL+1
LAST_COL=TARGET_COL

Sn=Side(style="thin"); Sd=Side(style="double"); S0=Side(style=None)
def bdr(l=S0,r=S0,t=S0,b=S0): return Border(left=l,right=r,top=t,bottom=b)
def fill(c): return PatternFill(start_color=c,end_color=c,fill_type="solid")

ph_fill=fill(COLOR_PH); jm_fill=fill(COLOR_JM); ot_fill=fill(COLOR_OT)
lb_fill=fill(COLOR_LABEL); yel_fill=fill(COLOR_YEL)

tf =Font(name="Meiryo",size=14,bold=True)
sf =Font(name="Meiryo",size=11,bold=True,color="FFFFFF")
nf =Font(name="Meiryo",size=10,bold=True)
lf =Font(name="Meiryo",size=9)
df =Font(name="Meiryo",size=9)
bf =Font(name="Meiryo",size=9,bold=True)
hf =Font(name="Meiryo",size=9,color="CC0000")
satf=Font(name="Meiryo",size=9,color="0000CC")
ctr =Alignment(horizontal="center",vertical="center")
lctr=Alignment(horizontal="left",  vertical="center")

def dfill(d):
    if is_sun_hol(d): return fill(COLOR_HOL)
    if datetime.date(YEAR,MONTH,d).weekday()==5: return fill(COLOR_SAT)
    return None
def dfont(d):
    if is_sun_hol(d): return hf
    if datetime.date(YEAR,MONTH,d).weekday()==5: return satf
    return df
def totime(s):
    h,m=s.split(":"); return datetime.time(int(h),int(m))

# ========== Row 1: タイトル ==========
reiwa=YEAR-2018
ws.merge_cells(start_row=1,start_column=COL_NAME,end_row=1,end_column=LAST_COL)
c=ws.cell(1,COL_NAME,f"大津京こと薬局　シフト表　令和{reiwa}年{MONTH}月")
c.font=tf; c.alignment=lctr
ws.row_dimensions[1].height=38.25
ws.row_dimensions[2].height=4.5

# ========== Row 3-4: 日付ヘッダー ==========
for d in range(1,DAYS_IN_MONTH+1):
    col=DAY_COL_START+d-1; fl=dfill(d); fn=dfont(d)
    c=ws.cell(3,col,d); c.font=fn; c.alignment=ctr
    c.border=bdr(l=Sn,r=Sn,t=Sn)
    if fl: c.fill=fl
    wd=datetime.date(YEAR,MONTH,d).weekday()
    lbl="祝" if d in HOLIDAYS else WEEKDAY_NAMES[wd]
    c=ws.cell(4,col,lbl); c.font=fn; c.alignment=ctr
    c.border=bdr(l=Sn,r=Sn,b=Sn)
    if fl: c.fill=fl

ws.merge_cells(start_row=3,start_column=TOTAL_COL,end_row=4,end_column=TOTAL_COL)
c=ws.cell(3,TOTAL_COL,"合計"); c.font=bf; c.alignment=ctr
c.border=bdr(l=Sd,r=Sn,t=Sn,b=Sn)
c=ws.cell(3,TARGET_COL,"目標"); c.font=df; c.alignment=ctr
c.border=bdr(l=Sn,r=Sn,t=Sn)
c=ws.cell(4,TARGET_COL,"差"); c.font=df; c.alignment=ctr
c.border=bdr(l=Sn,r=Sn,b=Sn)
ws.row_dimensions[3].height=15.75; ws.row_dimensions[4].height=15.75

def write_section(row,label,sec_fill):
    ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=LAST_COL)
    c=ws.cell(row,1,label); c.font=sf; c.fill=sec_fill; c.alignment=lctr
    ws.row_dimensions[row].height=18

def write_staff(start_row,sec_fill,key,target_h):
    data=shift_data[key]; off=requested_off[key]; yku=yukyu_per_person[key]
    labels=["入","退","勤務時間","備考"]
    or_=start_row; rr_=start_row+1; wr_=start_row+2
    for ro in range(4):
        r=start_row+ro; ts=Sn if ro==0 else S0; bs=Sn if ro==3 else S0
        ws.cell(r,COL_STRIPE).fill=sec_fill
        ws.cell(r,COL_STRIPE).alignment=lctr
        c=ws.cell(r,COL_NAME); c.alignment=ctr
        c.border=bdr(l=Sn,r=Sn,t=ts,b=bs)
        if ro==0: c.value=NAMES[key]; c.font=nf
        c=ws.cell(r,COL_TYPE,labels[ro]); c.font=lf; c.fill=lb_fill
        c.alignment=ctr; c.border=bdr(l=Sn,r=Sn,t=ts,b=bs)
        for d in range(1,DAYS_IN_MONTH+1):
            col=DAY_COL_START+d-1; cl=get_column_letter(col)
            c=ws.cell(r,col); c.alignment=ctr; c.font=df
            c.border=bdr(l=Sn,r=Sn,t=ts,b=bs)
            fl=dfill(d)
            if d in data:
                st,en,_,_=data[d]
                if fl: c.fill=fl
                if ro==0: c.value=totime(st); c.number_format="h:mm"
                elif ro==1: c.value=totime(en); c.number_format="h:mm"
                elif ro==2:
                    c.value=(f'=IF({cl}{or_}=0,"",({cl}{rr_}-{cl}{or_})*24'
                             f'-IF(({cl}{rr_}-{cl}{or_})*24>=7,1,'
                             f'IF(({cl}{rr_}-{cl}{or_})*24>6,0.5,0)))')
                    c.number_format="0.##"
            else:
                if fl: c.fill=fl
                if ro==3:
                    if d in yku: c.value="有休"
                    elif d in off: c.value="×"
        c=ws.cell(r,TOTAL_COL); c.alignment=ctr
        c.border=bdr(l=Sd,r=Sn,t=ts,b=bs)
        if ro==2:
            fc=get_column_letter(DAY_COL_START); lc=get_column_letter(DAY_COL_START+DAYS_IN_MONTH-1)
            c.value=f'=SUM({fc}{wr_}:{lc}{wr_})'; c.font=bf
        c=ws.cell(r,TARGET_COL); c.alignment=ctr
        c.border=bdr(l=Sn,r=Sn,t=ts,b=bs)
        if target_h is not None:
            if ro==2: c.value=target_h; c.font=df
            elif ro==3:
                tc=get_column_letter(TOTAL_COL); gc=get_column_letter(TARGET_COL)
                c.value=f'={tc}{wr_}-{gc}{wr_}'; c.font=df
    for i in range(3): ws.row_dimensions[start_row+i].height=15
    ws.row_dimensions[start_row+3].height=13.5

# ========== セクション配置 ==========
write_section(5,"【薬剤師】",ph_fill)
row=6
for k in ("A","B","C","D","E","F","G"):
    write_staff(row,ph_fill,k,TARGETS[k]); row+=4

write_section(row,"【医療事務】",jm_fill); row+=1
for k in ("H","I","J"):
    write_staff(row,jm_fill,k,TARGETS[k]); row+=4

write_section(row,"【その他】",ot_fill); row+=1
write_staff(row,ot_fill,"K",TARGETS["K"])

# ========== 列幅 ==========
ws.column_dimensions[get_column_letter(COL_STRIPE)].width=1.71
ws.column_dimensions[get_column_letter(COL_NAME)].width=14.14
ws.column_dimensions[get_column_letter(COL_TYPE)].width=6.0
for d in range(1,DAYS_IN_MONTH+1):
    ws.column_dimensions[get_column_letter(DAY_COL_START+d-1)].width=6.43
ws.column_dimensions[get_column_letter(TOTAL_COL)].width=7.0
ws.column_dimensions[get_column_letter(TARGET_COL)].width=7.0

# ================================================================
# 保存
# ================================================================
out_dir  = r"C:\Users\mibuk\OneDrive\デスクトップ\VSCode\大津京こと薬局"
out_path = os.path.join(out_dir, f"{YEAR}年{MONTH}月シフト.xlsx")
os.makedirs(out_dir, exist_ok=True)
wb.save(out_path)
print(f"保存完了: {out_path}")

# ================================================================
# 検証
# ================================================================
print(f"\n=== {YEAR}年{MONTH}月 シフト検証 ===")

print("\n--- 日祝担当ローテーション ---")
for d in sun_hol_days:
    lbl="祝" if d in HOLIDAYS else "日"
    wd=WEEKDAY_NAMES[datetime.date(YEAR,MONTH,d).weekday()]
    ph=ph_assign.get(d,"-"); jm=jm_assign.get(d,"-")
    comp=""
    if datetime.date(YEAR,MONTH,d).weekday()==6:
        sat=d-1
        if 1<=sat<=DAYS_IN_MONTH:
            comp=f" → {NAMES[ph]}・{NAMES[jm]} 土{sat}日オフ"
    print(f"  {d:2d}日({wd}/{lbl}): 薬={NAMES[ph]:9s} 事={NAMES[jm]:9s}{comp}")

print("\n--- 労働時間 ---")
for k in PERSONS:
    if not shift_data[k]: continue
    total=sum(v[2] for v in shift_data[k].values()); days=len(shift_data[k])
    tgt=TARGETS[k]
    if tgt:
        diff=total-tgt; flag="!" if abs(diff)>5 else "OK"
        print(f"{NAMES[k]:9s}: {total:6.2f}h / 目標{tgt:6.2f}h (差{diff:+.2f}h) {days:2d}日 {flag}")
    else:
        print(f"{NAMES[k]:9s}: {total:6.2f}h (目標未設定) {days:2d}日")

print("\n--- 連勤チェック ---")
for k in PERSONS:
    days_list=sorted(d for d in shift_data[k] if d not in yukyu_per_person[k])
    if not days_list: continue
    mx=streak=1
    for i in range(1,len(days_list)):
        streak=streak+1 if days_list[i]==days_list[i-1]+1 else 1
        mx=max(mx,streak)
    print(f"{NAMES[k]:9s}: 最大{mx}連勤{'  ← !要確認!' if mx>=6 else ''}")

print("\n--- 薬剤師カバレッジチェック ---")
cov_ok = True
rules = {0:"月(B+C)", 1:"火(B+D)", 2:"水(C必須)", 3:"木(全員)", 4:"金(2人)", 5:"土(2人)"}
for d in range(1, DAYS_IN_MONTH+1):
    wd = datetime.date(YEAR, MONTH, d).weekday()
    if is_sun_hol(d) or wd == 6: continue
    ph = [k for k in ("B","C","D") if d in shift_data[k]]
    cnt = len(ph)
    issue = None
    if wd == 0 and cnt < 2:   issue = f"B/C/D {cnt}人(要2)"
    elif wd == 1 and cnt < 2: issue = f"B/C/D {cnt}人(要2)"
    elif wd == 2 and "C" not in ph: issue = f"C不在"
    elif wd == 3 and cnt < 3: issue = f"B/C/D {cnt}人(要3)"
    elif wd in (4,5) and cnt < 2: issue = f"B/C/D {cnt}人(要2)"
    if issue:
        print(f"  {d:2d}日({WEEKDAY_NAMES[wd]}) NG: {issue}")
        cov_ok = False
if cov_ok:
    print("  全日OK")

print("\n--- 日祝配置チェック ---")
ph_k=("A","B","C","D","E","F","G"); jm_k=("H","I","J")
all_ok=True
for d in sun_hol_days:
    ph=[k for k in ph_k if d in shift_data[k]]
    jm=[k for k in jm_k if d in shift_data[k]]
    lbl="祝" if d in HOLIDAYS else "日"
    if len(ph)!=1 or len(jm)!=1:
        issues=[]
        if len(ph)!=1: issues.append(f"薬剤師{len(ph)}人(要1)")
        if len(jm)!=1: issues.append(f"事務{len(jm)}人(要1)")
        print(f"  {d:2d}日({lbl}) NG: {', '.join(issues)}")
        all_ok=False
    else:
        print(f"  {d:2d}日({lbl}) OK: {NAMES[ph[0]]} / {NAMES[jm[0]]}")
if all_ok:
    print("  → 全日祝OK")
