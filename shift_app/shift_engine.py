"""大津京こと薬局 シフト生成エンジン"""
import calendar
import datetime
import io

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins

WEEKDAY_NAMES = {0: "月", 1: "火", 2: "水", 3: "木", 4: "金", 5: "土", 6: "日"}

NAMES = {
    "A": "宮﨑 恵未",  "B": "中本 貴士",  "C": "今堀 翔太",
    "D": "安井 桜",    "E": "石田 真澄",  "F": "筒井 麻耶",
    "G": "黒田 梨恵",  "H": "西山 美紗子","I": "原田 真衣",
    "J": "和田 美聖",  "K": "小寺 和彦",
}
PERSONS = list("ABCDEFGHIJK")

COLOR_PH    = "4BACC6"
COLOR_JM    = "F79646"
COLOR_OT    = "70AD47"
COLOR_SAT   = "DDEEFF"
COLOR_HOL   = "F2DCDB"
COLOR_LABEL = "F2F2F2"
COLOR_YEL   = "FFFF00"

SHOTEIKOJI = {
    2026: {1:151, 2:143, 3:169, 4:163, 5:153,
           6:171, 7:169, 8:169, 9:147, 10:169, 11:155, 12:165},
}

PH_FIXED_OFF_WD = {"B": 2, "C": 1, "D": 0}


def net_hours(s, e):
    h1, m1 = map(int, s.split(":"))
    h2, m2 = map(int, e.split(":"))
    g = (h2 * 60 + m2 - h1 * 60 - m1) / 60
    return g - 1.0 if g > 6 else g


def to_time(s):
    h, m = s.split(":")
    return datetime.time(int(h), int(m))


def generate_shift(config: dict):
    year  = config["year"]
    month = config["month"]
    hols  = config.get("holidays", set())

    sun_override_ph = config.get("sun_override_ph", {})
    sun_override_jm = config.get("sun_override_jm", {})

    b_fri = set(config.get("b_fri", set()))
    c_fri = set(config.get("c_fri", set()))
    d_fri = set(config.get("d_fri", set()))
    b_sat = set(config.get("b_sat", set()))
    c_sat = set(config.get("c_sat", set()))
    d_sat = set(config.get("d_sat", set()))

    extra_off = {
        "B": set(config.get("b_extra_off", set())),
        "C": set(config.get("c_extra_off", set())),
        "D": set(config.get("d_extra_off", set())),
        "H": set(config.get("h_extra_off", set())),
        "I": set(config.get("i_extra_off", set())),
        "J": set(config.get("j_extra_off", set())),
    }

    requested_off    = config.get("requested_off",    {k: set() for k in PERSONS})
    yukyu_per_person = config.get("yukyu_per_person", {k: set() for k in PERSONS})

    days_in_month = calendar.monthrange(year, month)[1]

    def is_sun_hol(d):
        return datetime.date(year, month, d).weekday() == 6 or d in hols

    def skip(name, d):
        return d in requested_off.get(name, set()) or d in yukyu_per_person.get(name, set())

    weds  = [d for d in range(1, days_in_month + 1)
             if datetime.date(year, month, d).weekday() == 2]
    h_wed = set(weds[:2])
    i_wed = set(weds[2:])

    def sat_has_wed(sat_d, wed_set):
        wed_day = sat_d - 3
        return 1 <= wed_day <= days_in_month and wed_day in wed_set

    # ---- 日祝ローテーション ----
    sun_hol_days = sorted(d for d in range(1, days_in_month + 1) if is_sun_hol(d))
    ph_rot = ["B", "C", "D"]; jm_rot = ["H", "I", "J"]
    ph_assign = {}; jm_assign = {}
    ph_idx = jm_idx = 0

    for d in sun_hol_days:
        wd = datetime.date(year, month, d).weekday()
        if d in sun_override_ph:
            ph_assign[d] = sun_override_ph[d]
        else:
            for offset in range(3):
                candidate = ph_rot[(ph_idx + offset) % 3]
                if PH_FIXED_OFF_WD.get(candidate) != wd:
                    ph_assign[d] = candidate
                    ph_idx = (ph_idx + offset + 1) % 3
                    break
            else:
                ph_assign[d] = ph_rot[ph_idx % 3]
                ph_idx = (ph_idx + 1) % 3

        if d in sun_override_jm:
            jm_assign[d] = sun_override_jm[d]
        else:
            jm_assign[d] = jm_rot[jm_idx % 3]
            jm_idx = (jm_idx + 1) % 3

    # ---- 日曜補償日（自動追加）----
    for d, person in list(ph_assign.items()) + list(jm_assign.items()):
        if datetime.date(year, month, d).weekday() == 6 and person in extra_off:
            comp_d = (d + 1) if person in ("H", "I") else (d - 1)
            if 1 <= comp_d <= days_in_month:
                extra_off[person].add(comp_d)

    # ---- H：第1・2水曜の週の月曜を自動休みに ----
    for d in range(1, days_in_month + 1):
        if datetime.date(year, month, d).weekday() == 0:
            wed_of_week = d + 2
            if wed_of_week <= days_in_month and wed_of_week in h_wed:
                extra_off["H"].add(d)

    # ---- 金土自動配置（未指定の場合のみ）----
    if not any([b_fri, c_fri, d_fri, b_sat, c_sat, d_sat]):
        shoteikoji = SHOTEIKOJI.get(year, {}).get(month, 0)
        if shoteikoji:
            base = {"B": 0.0, "C": 0.0, "D": 0.0}
            for _d in range(1, days_in_month + 1):
                _wd = datetime.date(year, month, _d).weekday()
                _sh = is_sun_hol(_d)
                for _p in ("B", "C", "D"):
                    if _sh:
                        if ph_assign.get(_d) == _p: base[_p] += 7.0
                    else:
                        if _p == "B" and _wd in (0, 1, 3): base["B"] += 8.0
                        elif _p == "C":
                            if _wd in (0, 3):  base["C"] += 8.0
                            elif _wd == 2:     base["C"] += 7.0
                        elif _p == "D":
                            if _wd in (1, 3):  base["D"] += 8.0
                            elif _wd == 2:     base["D"] += 7.0

            rem  = {_p: max(0.0, shoteikoji - base[_p]) for _p in ("B", "C", "D")}
            fset = {"B": b_fri, "C": c_fri, "D": d_fri}
            sset = {"B": b_sat, "C": c_sat, "D": d_sat}

            for _d in sorted(d for d in range(1, days_in_month + 1)
                             if datetime.date(year, month, d).weekday() == 5 and not is_sun_hol(d)):
                avail = [_p for _p in ("B", "C", "D") if _d not in extra_off[_p]]
                for _p in sorted(avail, key=lambda p: -rem[p])[:2]:
                    sset[_p].add(_d); rem[_p] -= 7.0

            for _d in sorted(d for d in range(1, days_in_month + 1)
                             if datetime.date(year, month, d).weekday() == 4 and not is_sun_hol(d)):
                for _p in sorted(("B", "C", "D"), key=lambda p: -rem[p])[:2]:
                    fset[_p].add(_d); rem[_p] -= 8.0

    # ---- シフトデータ構築 ----
    shift_data = {n: {} for n in PERSONS}

    for d in range(1, days_in_month + 1):
        wd = datetime.date(year, month, d).weekday()
        sh = is_sun_hol(d)

        # A: 月火木金 8:45-17:30、水 9:00-17:30
        if not skip("A", d) and not sh:
            if wd in (0, 1, 3, 4):
                shift_data["A"][d] = ("8:45", "17:30", net_hours("8:45", "17:30"), False)
            elif wd == 2:
                shift_data["A"][d] = ("9:00", "17:00", net_hours("9:00", "17:00"), False)

        # B: 月遅・火遅・木早、水固定休
        if not skip("B", d) and wd != 2 and d not in extra_off["B"]:
            if sh:
                if ph_assign.get(d) == "B":
                    shift_data["B"][d] = ("9:00", "17:00", net_hours("9:00", "17:00"), False)
            elif wd == 0: shift_data["B"][d] = ("10:00", "19:00", net_hours("10:00", "19:00"), False)
            elif wd == 1: shift_data["B"][d] = ("10:00", "19:00", net_hours("10:00", "19:00"), False)
            elif wd == 3: shift_data["B"][d] = ("9:00",  "18:00", net_hours("9:00",  "18:00"), False)
            elif wd == 4 and d in b_fri:
                shift_data["B"][d] = ("10:00", "19:00", net_hours("10:00", "19:00"), False)
            elif wd == 5 and d in b_sat:
                shift_data["B"][d] = ("9:00",  "17:00", net_hours("9:00",  "17:00"), False)

        # C: 月遅・水9-17・木遅、火固定休
        if not skip("C", d) and wd != 1 and d not in extra_off["C"]:
            if sh:
                if ph_assign.get(d) == "C":
                    shift_data["C"][d] = ("9:00", "17:00", net_hours("9:00", "17:00"), False)
            elif wd == 0: shift_data["C"][d] = ("10:00", "19:00", net_hours("10:00", "19:00"), False)
            elif wd == 2: shift_data["C"][d] = ("9:00",  "17:00", net_hours("9:00",  "17:00"), False)
            elif wd == 3: shift_data["C"][d] = ("10:00", "19:00", net_hours("10:00", "19:00"), False)
            elif wd == 4 and d in c_fri:
                shift_data["C"][d] = ("10:00", "19:00", net_hours("10:00", "19:00"), False)
            elif wd == 5 and d in c_sat:
                shift_data["C"][d] = ("9:00",  "17:00", net_hours("9:00",  "17:00"), False)

        # D: 火遅・水9-17・木遅、月固定休
        if not skip("D", d) and wd != 0 and d not in extra_off["D"]:
            if sh:
                if ph_assign.get(d) == "D":
                    shift_data["D"][d] = ("9:00", "17:00", net_hours("9:00", "17:00"), False)
            elif wd == 1: shift_data["D"][d] = ("10:00", "19:00", net_hours("10:00", "19:00"), False)
            elif wd == 2: shift_data["D"][d] = ("9:00",  "17:00", net_hours("9:00",  "17:00"), False)
            elif wd == 3: shift_data["D"][d] = ("10:00", "19:00", net_hours("10:00", "19:00"), False)
            elif wd == 4 and d in d_fri:
                shift_data["D"][d] = ("10:00", "19:00", net_hours("10:00", "19:00"), False)
            elif wd == 5 and d in d_sat:
                shift_data["D"][d] = ("9:00",  "17:00", net_hours("9:00",  "17:00"), False)

        # E: 月火木金 9:00-13:00
        if not skip("E", d) and not sh and wd in (0, 1, 3, 4):
            shift_data["E"][d] = ("9:00", "13:00", net_hours("9:00", "13:00"), False)

        # F: 月火金 9:00-14:00
        if not skip("F", d) and not sh and wd in (0, 1, 4):
            shift_data["F"][d] = ("9:00", "14:00", net_hours("9:00", "14:00"), False)

        # G: 木 9:00-13:00、土 9:00-17:00
        if not skip("G", d):
            if wd == 3 and not sh:
                shift_data["G"][d] = ("9:00", "13:00", net_hours("9:00", "13:00"), False)
            elif wd == 5:
                shift_data["G"][d] = ("9:00", "17:00", net_hours("9:00", "17:00"), False)

        # H: 月遅・火早・第1,2水 8:45-17・木遅・金早
        if not skip("H", d) and d not in extra_off["H"]:
            if sh:
                if jm_assign.get(d) == "H":
                    shift_data["H"][d] = ("9:00", "17:00", net_hours("9:00", "17:00"), False)
            elif d in h_wed:
                shift_data["H"][d] = ("9:00", "17:00", net_hours("9:00", "17:00"), False)
            elif wd == 0: shift_data["H"][d] = ("10:00", "19:00", net_hours("10:00", "19:00"), False)
            elif wd == 1: shift_data["H"][d] = ("8:45",  "18:00", net_hours("8:45",  "18:00"), False)
            elif wd == 3: shift_data["H"][d] = ("10:00", "19:00", net_hours("10:00", "19:00"), False)
            elif wd == 4: shift_data["H"][d] = ("8:45",  "18:00", net_hours("8:45",  "18:00"), False)
            elif wd == 5 and not sat_has_wed(d, h_wed):
                shift_data["H"][d] = ("9:00", "17:00", net_hours("9:00", "17:00"), False)

        # I: 月早・火遅・第3,4水 8:45-17・木早・金遅
        if not skip("I", d) and d not in extra_off["I"]:
            if sh:
                if jm_assign.get(d) == "I":
                    shift_data["I"][d] = ("9:00", "17:00", net_hours("9:00", "17:00"), False)
            elif d in i_wed:
                shift_data["I"][d] = ("9:00", "17:00", net_hours("9:00", "17:00"), False)
            elif wd == 0: shift_data["I"][d] = ("8:45",  "18:00", net_hours("8:45",  "18:00"), False)
            elif wd == 1: shift_data["I"][d] = ("10:00", "19:00", net_hours("10:00", "19:00"), False)
            elif wd == 3: shift_data["I"][d] = ("8:45",  "18:00", net_hours("8:45",  "18:00"), False)
            elif wd == 4: shift_data["I"][d] = ("10:00", "19:00", net_hours("10:00", "19:00"), False)
            elif wd == 5 and not sat_has_wed(d, i_wed):
                shift_data["I"][d] = ("9:00", "17:00", net_hours("9:00", "17:00"), False)

        # J: 月早・火遅・木遅・金早・土 9-17、水固定休
        if not skip("J", d) and wd != 2 and d not in extra_off["J"]:
            if sh:
                if jm_assign.get(d) == "J":
                    shift_data["J"][d] = ("9:00", "17:00", net_hours("9:00", "17:00"), False)
            elif wd == 0: shift_data["J"][d] = ("8:45",  "18:00", net_hours("8:45",  "18:00"), False)
            elif wd == 1: shift_data["J"][d] = ("10:00", "19:00", net_hours("10:00", "19:00"), False)
            elif wd == 3: shift_data["J"][d] = ("10:00", "19:00", net_hours("10:00", "19:00"), False)
            elif wd == 4: shift_data["J"][d] = ("8:45",  "18:00", net_hours("8:45",  "18:00"), False)
            elif wd == 5: shift_data["J"][d] = ("9:00",  "17:00", net_hours("9:00",  "17:00"), False)

    # ---- 目標時間 ----
    shoteikoji = SHOTEIKOJI.get(year, {}).get(month)
    targets = {}
    for n in PERSONS:
        if n == "A":
            targets[n] = round(shoteikoji * 0.8, 2) if shoteikoji else round(sum(v[2] for v in shift_data[n].values()), 2)
        elif n in ("E", "F", "G", "K"):
            targets[n] = round(sum(v[2] for v in shift_data[n].values()), 2)
        else:
            targets[n] = shoteikoji

    verif = _verify(shift_data, targets, yukyu_per_person, requested_off,
                    days_in_month, year, month, hols, is_sun_hol,
                    ph_assign, jm_assign, sun_hol_days)
    verif["ph_assign"]     = ph_assign
    verif["jm_assign"]     = jm_assign
    verif["sun_hol_days"]  = sun_hol_days

    excel = _build_excel(shift_data, targets, requested_off, yukyu_per_person,
                         year, month, days_in_month, hols, is_sun_hol)

    return shift_data, excel, verif


def _verify(shift_data, targets, yukyu, r_off, days_in_month, year, month,
            hols, is_sun_hol, ph_assign, jm_assign, sun_hol_days):
    results = {"hours": {}, "streaks": {}, "coverage_issues": [], "sunhol_issues": []}

    for n in PERSONS:
        if not shift_data[n]:
            continue
        total  = sum(v[2] for v in shift_data[n].values())
        target = targets.get(n)
        diff   = (total - target) if target else 0.0
        results["hours"][n] = {"total": total, "target": target, "diff": diff}

    for n in PERSONS:
        yu = yukyu.get(n, set())
        dl = sorted(d for d in shift_data[n] if d not in yu)
        mx = streak = (1 if dl else 0)
        for i in range(1, len(dl)):
            if dl[i] == dl[i - 1] + 1:
                streak += 1; mx = max(mx, streak)
            else:
                streak = 1
        results["streaks"][n] = mx

    for d in range(1, days_in_month + 1):
        wd = datetime.date(year, month, d).weekday()
        if is_sun_hol(d) or wd == 6:
            continue
        ph = [k for k in ("B", "C", "D")
              if d in shift_data[k] and d not in yukyu.get(k, set())]
        cnt = len(ph)
        issue = None
        if wd == 0 and cnt < 2:            issue = f"月(B/C/D) {cnt}人（要2）"
        elif wd == 1 and cnt < 2:          issue = f"火(B/C/D) {cnt}人（要2）"
        elif wd == 2 and "C" not in ph:    issue = "水(C不在)"
        elif wd == 3 and cnt < 3:          issue = f"木(B/C/D) {cnt}人（要3）"
        elif wd == 4 and cnt < 2:          issue = f"金(B/C/D) {cnt}人（要2）"
        elif wd == 5 and cnt < 2:          issue = f"土(B/C/D) {cnt}人（要2）"
        if issue:
            results["coverage_issues"].append(
                f"{d}日（{WEEKDAY_NAMES[wd]}）NG: {issue}")

    ph_k = ("A", "B", "C", "D", "E", "F", "G")
    jm_k = ("H", "I", "J")
    for d in sun_hol_days:
        ph_cnt = [k for k in ph_k if d in shift_data[k]]
        jm_cnt = [k for k in jm_k if d in shift_data[k]]
        lbl = "祝" if d in hols else "日"
        issues = []
        if len(ph_cnt) != 1: issues.append(f"薬剤師{len(ph_cnt)}人（要1）")
        if len(jm_cnt) != 1: issues.append(f"事務{len(jm_cnt)}人（要1）")
        if issues:
            results["sunhol_issues"].append(
                f"{d}日（{lbl}）NG: {', '.join(issues)}")

    return results


def _build_excel(shift_data, targets, r_off, yukyu, year, month, days_in_month, hols, is_sun_hol):
    Sn = Side(style="thin"); Sd = Side(style="double"); S0 = Side(style=None)

    def bdr(l=S0, r=S0, t=S0, b=S0):
        return Border(left=l, right=r, top=t, bottom=b)

    def fill(c):
        return PatternFill(start_color=c, end_color=c, fill_type="solid")

    ph_fill  = fill(COLOR_PH);  jm_fill = fill(COLOR_JM)
    ot_fill  = fill(COLOR_OT);  lb_fill = fill(COLOR_LABEL)
    yel_fill = fill(COLOR_YEL)

    tf   = Font(name="Meiryo", size=14, bold=True)
    sf   = Font(name="Meiryo", size=11, bold=True, color="FFFFFF")
    nf   = Font(name="Meiryo", size=10, bold=True)
    lf   = Font(name="Meiryo", size=9)
    df   = Font(name="Meiryo", size=9)
    bf   = Font(name="Meiryo", size=9, bold=True)
    hf   = Font(name="Meiryo", size=9, color="CC0000")
    satf = Font(name="Meiryo", size=9, color="0000CC")
    ctr  = Alignment(horizontal="center", vertical="center")
    lctr = Alignment(horizontal="left",   vertical="center")

    def dfill(d):
        if is_sun_hol(d):                               return fill(COLOR_HOL)
        if datetime.date(year, month, d).weekday() == 5: return fill(COLOR_SAT)
        return None

    def dfont(d):
        if is_sun_hol(d):                               return hf
        if datetime.date(year, month, d).weekday() == 5: return satf
        return df

    reiwa = year - 2018
    wb = Workbook(); ws = wb.active
    ws.title = "薬局シフト表"
    ws.page_setup.paperSize   = ws.PAPERSIZE_A4
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins = PageMargins(
        left=0.7, right=0.7, top=0.75, bottom=0.75, header=0.3, footer=0.3)

    CS = 1; CN = 2; CT = 3; DC = 4
    TC = DC + days_in_month; GC = TC + 1

    ws.merge_cells(start_row=1, start_column=CN, end_row=1, end_column=GC)
    c = ws.cell(1, CN, f"大津京こと薬局　シフト表　令和{reiwa}年{month}月")
    c.font = tf; c.alignment = lctr
    ws.row_dimensions[1].height = 38.25
    ws.row_dimensions[2].height = 4.5

    for d in range(1, days_in_month + 1):
        col = DC + d - 1; fl = dfill(d); fn = dfont(d)
        c = ws.cell(3, col, d); c.font = fn; c.alignment = ctr
        c.border = bdr(l=Sn, r=Sn, t=Sn)
        if fl: c.fill = fl
        wd  = datetime.date(year, month, d).weekday()
        lbl = "祝" if d in hols else WEEKDAY_NAMES[wd]
        c = ws.cell(4, col, lbl); c.font = fn; c.alignment = ctr
        c.border = bdr(l=Sn, r=Sn, b=Sn)
        if fl: c.fill = fl

    ws.merge_cells(start_row=3, start_column=TC, end_row=4, end_column=TC)
    c = ws.cell(3, TC, "合計"); c.font = bf; c.alignment = ctr
    c.border = bdr(l=Sd, r=Sn, t=Sn, b=Sn)
    c = ws.cell(3, GC, "目標"); c.font = df; c.alignment = ctr
    c.border = bdr(l=Sn, r=Sn, t=Sn)
    c = ws.cell(4, GC, "差");   c.font = df; c.alignment = ctr
    c.border = bdr(l=Sn, r=Sn, b=Sn)
    ws.row_dimensions[3].height = 15.75; ws.row_dimensions[4].height = 15.75

    def write_section(row, label, sec_fill):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=GC)
        c = ws.cell(row, 1, label); c.font = sf; c.fill = sec_fill; c.alignment = lctr
        ws.row_dimensions[row].height = 18

    def write_staff(start_row, sec_fill, key, target_h):
        data = shift_data[key]
        off  = r_off.get(key, set())
        yku  = yukyu.get(key, set())
        lbls = ["入", "退", "勤務時間", "備考"]
        or_  = start_row; rr_ = start_row + 1; wr_ = start_row + 2

        for ro in range(4):
            r  = start_row + ro
            ts = Sn if ro == 0 else S0
            bs = Sn if ro == 3 else S0
            ws.cell(r, CS).fill = sec_fill; ws.cell(r, CS).alignment = lctr
            c = ws.cell(r, CN); c.alignment = ctr
            c.border = bdr(l=Sn, r=Sn, t=ts, b=bs)
            if ro == 0: c.value = NAMES[key]; c.font = nf
            c = ws.cell(r, CT, lbls[ro]); c.font = lf; c.fill = lb_fill
            c.alignment = ctr; c.border = bdr(l=Sn, r=Sn, t=ts, b=bs)

            for d in range(1, days_in_month + 1):
                col = DC + d - 1; cl = get_column_letter(col)
                c = ws.cell(r, col); c.alignment = ctr; c.font = df
                c.border = bdr(l=Sn, r=Sn, t=ts, b=bs)
                fl = dfill(d)
                if d in data:
                    st, en, _, adj = data[d]
                    if adj: c.fill = yel_fill
                    elif fl: c.fill = fl
                    if ro == 0:
                        c.value = to_time(st); c.number_format = "h:mm"
                    elif ro == 1:
                        c.value = to_time(en); c.number_format = "h:mm"
                    elif ro == 2:
                        c.value = (f'=IF({cl}{or_}=0,"",({cl}{rr_}-{cl}{or_})*24'
                                   f'-IF(({cl}{rr_}-{cl}{or_})*24>6,1,0))')
                        c.number_format = "0.##"
                else:
                    if fl: c.fill = fl
                    if ro == 3:
                        if d in yku: c.value = "有休"
                        elif d in off: c.value = "×"

            c = ws.cell(r, TC); c.alignment = ctr
            c.border = bdr(l=Sd, r=Sn, t=ts, b=bs)
            if ro == 2:
                fc = get_column_letter(DC); lc = get_column_letter(DC + days_in_month - 1)
                c.value = f"=SUM({fc}{wr_}:{lc}{wr_})"; c.font = bf

            c = ws.cell(r, GC); c.alignment = ctr
            c.border = bdr(l=Sn, r=Sn, t=ts, b=bs)
            if target_h is not None:
                if ro == 2: c.value = target_h; c.font = df
                elif ro == 3:
                    c.value = f"={get_column_letter(TC)}{wr_}-{get_column_letter(GC)}{wr_}"
                    c.font = df

        for i in range(3): ws.row_dimensions[start_row + i].height = 15
        ws.row_dimensions[start_row + 3].height = 13.5

    write_section(5, "【薬剤師】", ph_fill)
    row = 6
    for k in ("A", "B", "C", "D", "E", "F", "G"):
        write_staff(row, ph_fill, k, targets[k]); row += 4

    write_section(row, "【医療事務】", jm_fill); row += 1
    for k in ("H", "I", "J"):
        write_staff(row, jm_fill, k, targets[k]); row += 4

    write_section(row, "【その他】", ot_fill); row += 1
    write_staff(row, ot_fill, "K", targets["K"])

    ws.column_dimensions[get_column_letter(CS)].width = 1.71
    ws.column_dimensions[get_column_letter(CN)].width = 14.14
    ws.column_dimensions[get_column_letter(CT)].width = 6.0
    for d in range(1, days_in_month + 1):
        ws.column_dimensions[get_column_letter(DC + d - 1)].width = 6.43
    ws.column_dimensions[get_column_letter(TC)].width = 7.0
    ws.column_dimensions[get_column_letter(GC)].width = 7.0

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return buf
